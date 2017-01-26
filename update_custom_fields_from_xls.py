from openpyxl import load_workbook
from redmine_rest_wrapper import custom_fields_as_payload
from redmine_rest import frw

filename = 'timelog1.xlsx'
wb = load_workbook(filename)
print wb.get_sheet_names()
ws = wb[wb.get_sheet_names()[0]]

issue_values = dict()

current_value = None

for i in range(1, 1024*1024):
    #print ws.cell(row=i, column=2).value
    bvalue = ws.cell(row=i, column=2).value
    if bvalue is None:
        pass
    else:
        project = ws.cell(row=i, column=1).value
        if project is None:
            # bvalue is an issue ID
            issue_values[current_value].append(bvalue)
        else:
            # bvalue is a "custom field value"
            current_value = bvalue
            if issue_values.get(current_value) is None:
                issue_values[current_value] = []

for value in issue_values:
    print value, issue_values[value]
    cf = custom_fields_as_payload({15: value})
    issues = set(issue_values[value])
    frw.put_issues_with_payload(issues, cf)
